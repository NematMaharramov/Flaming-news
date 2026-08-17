"""
Flaming News - Excel Writer
Rebuilds the workbook to match the ORIGINAL template's structure, merges,
colours and fonts exactly (extracted directly from the source .xlsx), filled
in with the day's data. No VIP Code Legend section (removed per requirement).
"""
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

import rules
import data_store

# ---------------------------------------------------------------------------
# Colours / styles lifted directly from the source template
# ---------------------------------------------------------------------------
GREEN = 'FF92D050'      # section / column header green
GREY = 'FFE7E6E6'       # value-cell background (theme "Background 2")
RED_TEXT = 'FFFF0000'   # title / quote text colour
WHITE = 'FFFFFFFF'
BLACK = 'FF000000'

FONT_NAME = 'Calibri'

THIN = Side(style='thin')
MEDIUM = Side(style='medium')

TITLE_FONT = Font(name=FONT_NAME, size=18, bold=True, color=RED_TEXT)
SECTION_FONT = Font(name=FONT_NAME, size=16, bold=True)
SECTION_FONT_PLAIN = Font(name=FONT_NAME, size=16, bold=False)
LABEL_FONT = Font(name=FONT_NAME, size=16, bold=True)
VALUE_FONT = Font(name=FONT_NAME, size=16, bold=False)
TABLE_HEADER_FONT = Font(name=FONT_NAME, size=16, bold=True)
DATA_FONT = Font(name=FONT_NAME, size=12, bold=False)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
CENTER_NOWRAP = Alignment(horizontal='center', vertical='center', wrap_text=False)

GREEN_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
GREY_FILL = PatternFill(start_color=GREY, end_color=GREY, fill_type='solid')


def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid') if color else None


def _border(left=None, right=None, top=None, bottom=None):
    return Border(left=left, right=right, top=top, bottom=bottom)


def _set(ws, coord, value=None, font=None, fill=None, align=None, border=None, numfmt=None):
    cell = ws[coord]
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    if numfmt:
        cell.number_format = numfmt
    return cell


def _merge(ws, rng):
    ws.merge_cells(rng)


def _vip_row(ws, row, guest, code, room, col_d_value, remark_or_time, company,
             fill_color=None, font_color=None, photo_fs_path=None, has_photo_col=False):
    """
    One VIP table row.
    Without photo column: Guest | Code | Room | D | Remarks/Time(E:F) | Company(G:I)
    With photo column:    Guest | Code | Room | D | Photo(E) | Remarks(F:G) | Company(H:I)
    """
    b = _border(THIN, THIN, THIN, THIN)
    _set(ws, f'A{row}', guest, DATA_FONT, None, CENTER, b)
    code_cell = _set(ws, f'B{row}', code, Font(name=FONT_NAME, size=12, bold=True), None, CENTER_NOWRAP, b)
    if fill_color:
        code_cell.fill = _fill(fill_color)
        code_cell.font = Font(name=FONT_NAME, size=12, bold=True, color=font_color or BLACK)
    _set(ws, f'C{row}', room, DATA_FONT, None, CENTER, b)
    _set(ws, f'D{row}', col_d_value, DATA_FONT, None, CENTER, b)

    if has_photo_col:
        _set(ws, f'E{row}', None, DATA_FONT, None, CENTER, b)
        if photo_fs_path:
            _embed_photo(ws, f'E{row}', photo_fs_path)
        _merge(ws, f'F{row}:G{row}')
        _set(ws, f'F{row}', remark_or_time, DATA_FONT, None, CENTER, b)
        _set(ws, f'G{row}', None, DATA_FONT, None, CENTER, b)
        _merge(ws, f'H{row}:I{row}')
        _set(ws, f'H{row}', company, DATA_FONT, None, CENTER, b)
        _set(ws, f'I{row}', None, DATA_FONT, None, CENTER, b)
        ws.row_dimensions[row].height = 32  # fixed row height — same for every guest, photo or not
    else:
        _merge(ws, f'E{row}:F{row}')
        _set(ws, f'E{row}', remark_or_time, DATA_FONT, None, CENTER, b)
        _set(ws, f'F{row}', None, DATA_FONT, None, CENTER, b)
        _merge(ws, f'G{row}:I{row}')
        _set(ws, f'G{row}', company, DATA_FONT, None, CENTER, b)
        for c in 'HI':
            _set(ws, f'{c}{row}', None, DATA_FONT, None, CENTER, b)


def _embed_photo(ws, coord, fs_path):
    """Insert a fixed-size (so every row stays the same height) guest photo.
    Uses an explicit OneCellAnchor with a fixed extent — setting
    Image.width/height alone does not reliably persist through save/reload
    in openpyxl, since add_image() can re-derive the size from the source
    file's native dimensions unless the anchor's extent is set explicitly."""
    try:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

        col_letter, row_num = coordinate_from_string(coord)
        col_idx = column_index_from_string(col_letter) - 1  # 0-based
        row_idx = row_num - 1  # 0-based

        img = XLImage(fs_path)
        size = XDRPositiveSize2D(pixels_to_EMU(28), pixels_to_EMU(28))
        marker = AnchorMarker(col=col_idx, colOff=pixels_to_EMU(2), row=row_idx, rowOff=pixels_to_EMU(2))
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(img)
    except Exception:
        pass  # missing/corrupt photo file — don't fail the whole export over it


def _photo_fs_path(photo_url):
    """'/photos/2026-08-16/abcd.jpg' -> absolute filesystem path, or None."""
    if not photo_url:
        return None
    rel = photo_url.lstrip('/')  # 'photos/2026-08-16/abcd.jpg'
    fs_path = os.path.join(data_store.DATA_DIR, rel)
    return fs_path if os.path.isfile(fs_path) else None


def _vip_table_header(ws, row, d_label, e_label, headerspan_top=True, has_photo_col=False):
    b = _border(THIN, THIN, MEDIUM if headerspan_top else THIN, MEDIUM)
    _set(ws, f'A{row}', 'Guest', TABLE_HEADER_FONT, None, CENTER, _border(MEDIUM, THIN, MEDIUM, MEDIUM))
    _set(ws, f'B{row}', 'Code', TABLE_HEADER_FONT, None, CENTER, _border(THIN, THIN, MEDIUM, MEDIUM))
    _set(ws, f'C{row}', 'Room', TABLE_HEADER_FONT, None, CENTER, _border(THIN, THIN, MEDIUM, MEDIUM))
    _set(ws, f'D{row}', d_label, TABLE_HEADER_FONT, GREY_FILL if d_label != 'ETA' else None, CENTER, _border(THIN, THIN, MEDIUM, MEDIUM))

    if has_photo_col:
        _set(ws, f'E{row}', 'Photo', TABLE_HEADER_FONT, None, CENTER, _border(THIN, THIN, MEDIUM, MEDIUM))
        _merge(ws, f'F{row}:G{row}')
        _set(ws, f'F{row}', e_label, TABLE_HEADER_FONT, None, CENTER, _border(THIN, THIN, MEDIUM, MEDIUM))
        _set(ws, f'G{row}', None, None, None, None, _border(THIN, THIN, MEDIUM, MEDIUM))
        _merge(ws, f'H{row}:I{row}')
        _set(ws, f'H{row}', 'Company', TABLE_HEADER_FONT, None, CENTER, _border(THIN, MEDIUM, MEDIUM, MEDIUM))
        _set(ws, f'I{row}', None, None, None, None, _border(THIN, MEDIUM, MEDIUM, MEDIUM))
    else:
        _merge(ws, f'E{row}:F{row}')
        _set(ws, f'E{row}', e_label, TABLE_HEADER_FONT, None, CENTER, _border(THIN, THIN, MEDIUM, MEDIUM))
        _set(ws, f'F{row}', None, None, None, None, _border(THIN, THIN, MEDIUM, MEDIUM))
        _merge(ws, f'G{row}:I{row}')
        _set(ws, f'G{row}', 'Company', TABLE_HEADER_FONT, None, CENTER, _border(THIN, MEDIUM, MEDIUM, MEDIUM))
        for c in 'HI':
            _set(ws, f'{c}{row}', None, None, None, None, _border(THIN, MEDIUM, MEDIUM, MEDIUM))


def _section_header(ws, row, text, span_end='I', full_border=True):
    _merge(ws, f'A{row}:{span_end}{row}')
    b = _border(MEDIUM, MEDIUM, MEDIUM, THIN) if full_border else None
    _set(ws, f'A{row}', text, Font(name=FONT_NAME, size=16, bold=True), GREEN_FILL, CENTER, b)
    for col in range(2, 10):
        cl = get_column_letter(col)
        if cl <= span_end:
            _set(ws, f'{cl}{row}', None, None, GREEN_FILL, None, b)


def build_workbook(data):
    """
    data: the day's JSON dict (see data_store.empty_day for schema).
    Returns an in-memory .xlsx (BytesIO).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Accomodation'

    report_date = data.get('iso_date', '')
    try:
        from datetime import datetime as _dt
        display_date = _dt.strptime(report_date, '%Y-%m-%d').strftime('%d %B %Y')
    except Exception:
        display_date = report_date
    title_text = f'FLAMING NEWS FOR {display_date}'
    quote_text = f"Quote of the Day : \u201c{data.get('quote', '')}\u201d"

    # ---- Row 1: Title / Row 2: Quote ----
    _merge(ws, 'A1:I1')
    _set(ws, 'A1', title_text, TITLE_FONT, GREY_FILL, CENTER, _border(THIN, THIN, THIN, THIN))
    _merge(ws, 'A2:I2')
    _set(ws, 'A2', quote_text, TITLE_FONT, GREY_FILL, CENTER, _border(THIN, THIN, THIN, MEDIUM))

    # ---- 7 Day Forecast ----
    _set(ws, 'A3', '7 day Forecast', SECTION_FONT, GREEN_FILL, CENTER, _border(MEDIUM, THIN, MEDIUM, THIN))
    _set(ws, 'B3', 'DATE:', SECTION_FONT, GREY_FILL, CENTER, _border(THIN, THIN, MEDIUM, THIN))
    fc = data.get('forecast', {})
    dates = fc.get('dates', [''] * 7)
    for i in range(7):
        col = get_column_letter(3 + i)
        _set(ws, f'{col}3', dates[i] if i < len(dates) else '', SECTION_FONT, GREY_FILL, CENTER,
             _border(THIN, MEDIUM if i == 6 else THIN, MEDIUM, THIN))

    metric_rows = [
        (4, 'Occupancy %', fc.get('occupancy_pct', [])),
        (5, 'Rooms Occupied\n(excl house use & comp)', fc.get('rooms_occupied', [])),
        (6, 'ADR', fc.get('adr', [])),
        (7, 'Arrivals', fc.get('arrivals', [])),
        (8, 'Departures ', fc.get('departures', [])),
    ]
    for row, label, values in metric_rows:
        _set(ws, f'A{row}', label, LABEL_FONT, GREEN_FILL, CENTER, _border(MEDIUM, THIN, THIN, THIN))
        for i in range(7):
            col = get_column_letter(3 + i)
            v = values[i] if i < len(values) and values[i] != '' else None
            _set(ws, f'{col}{row}', v, VALUE_FONT, GREY_FILL, CENTER,
                 _border(THIN, MEDIUM if i == 6 else THIN, THIN, THIN))

    # ---- AM MOD / PM MOD / NM / Weekend EOD  +  House Status / Weather / Enrollments ----
    _set(ws, 'A9', 'AM MOD:', LABEL_FONT, GREY_FILL, Alignment(horizontal='left', vertical='top', wrap_text=True), _border(MEDIUM, THIN, MEDIUM, THIN))
    _set(ws, 'B9', data.get('am_mod', ''), VALUE_FONT, GREY_FILL, None, _border(THIN, MEDIUM, MEDIUM, THIN))
    _set(ws, 'A10', 'PM MOD:', LABEL_FONT, GREY_FILL, Alignment(horizontal='left', vertical='top', wrap_text=True), _border(MEDIUM, THIN, THIN, THIN))
    _set(ws, 'B10', data.get('pm_mod', ''), VALUE_FONT, GREY_FILL, None, _border(THIN, MEDIUM, THIN, THIN))
    _set(ws, 'A11', 'NM:', LABEL_FONT, GREY_FILL, Alignment(horizontal='left', vertical='top', wrap_text=True), _border(MEDIUM, THIN, THIN, THIN))
    _set(ws, 'B11', data.get('nm', ''), VALUE_FONT, GREY_FILL, None, _border(THIN, MEDIUM, THIN, THIN))
    _set(ws, 'A12', 'Weekend EOD:', LABEL_FONT, GREY_FILL, Alignment(horizontal='left', vertical='top', wrap_text=True), _border(MEDIUM, THIN, THIN, MEDIUM))
    _set(ws, 'B12', data.get('weekend_eod', ''), VALUE_FONT, GREY_FILL, None, _border(THIN, MEDIUM, THIN, MEDIUM))

    hs = data.get('house_status', {})
    _merge(ws, 'C9:D9')
    _set(ws, 'C9', 'House Status', LABEL_FONT, GREEN_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, THIN))
    _set(ws, 'D9', None, None, GREEN_FILL, None, _border(MEDIUM, MEDIUM, MEDIUM, THIN))
    _set(ws, 'C10', 'OOS/ OOO', VALUE_FONT, None, CENTER, _border(THIN, THIN, THIN, THIN))
    _set(ws, 'D10', hs.get('oos_ooo', ''), VALUE_FONT, None, CENTER, _border(THIN, THIN, THIN, THIN))
    _set(ws, 'C11', 'No Show', VALUE_FONT, None, CENTER, _border(THIN, THIN, THIN, THIN))
    _set(ws, 'D11', hs.get('no_show', ''), VALUE_FONT, None, CENTER, _border(THIN, THIN, THIN, THIN))
    _set(ws, 'C12', 'Comp./house use', VALUE_FONT, None, CENTER, _border(THIN, THIN, THIN, MEDIUM))
    _set(ws, 'D12', hs.get('comp_house_use', ''), VALUE_FONT, None, CENTER, _border(THIN, THIN, THIN, MEDIUM))

    _merge(ws, 'E9:E10')
    _set(ws, 'E9', 'Weather Today', LABEL_FONT, GREEN_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, THIN))
    _merge(ws, 'E11:E12')
    _set(ws, 'E11', data.get('weather', ''), VALUE_FONT, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, THIN, MEDIUM))

    fg = data.get('fairmont_goals', {})
    _merge(ws, 'F9:G9')
    _set(ws, 'F9', 'ALL Enrollments Goal', LABEL_FONT, GREEN_FILL, CENTER, _border(MEDIUM, THIN, MEDIUM, THIN))
    _merge(ws, 'F10:G12')
    _set(ws, 'F10', data.get('enrollments_goal', ''), Font(name=FONT_NAME, size=20, bold=True), GREY_FILL, CENTER, _border(MEDIUM, THIN, THIN, MEDIUM))
    _merge(ws, 'H9:I9')
    _set(ws, 'H9', 'ALL Enrollments YTD', LABEL_FONT, GREEN_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, THIN))
    _merge(ws, 'H10:I12')
    _set(ws, 'H10', data.get('enrollments_ytd', ''), Font(name=FONT_NAME, size=20, bold=True), GREY_FILL, CENTER, _border(THIN, MEDIUM, THIN, MEDIUM))

    # ---- Fairmont Baku Goals ----
    _merge(ws, 'A13:I13')
    _set(ws, 'A13', 'Fairmont Baku 2026 Goals', Font(name=FONT_NAME, size=16), _fill('FF70AD47'), CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM))
    _set(ws, 'A14', 'CES', SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM))
    _set(ws, 'C14', fg.get('ces_goal', ''), SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM))
    _set(ws, 'D14', 'LQA', SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM))
    _set(ws, 'E14', fg.get('lqa_goal', ''), SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM))
    _set(ws, 'F14', 'RPS', SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM))
    _set(ws, 'A15', fg.get('ces_actual', ''), SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, None, MEDIUM, MEDIUM), numfmt='0%')
    _set(ws, 'C15', fg.get('ces_actual', ''), SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM), numfmt='0%')
    _set(ws, 'D15', fg.get('lqa_actual', ''), SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM))
    _set(ws, 'E15', fg.get('lqa_actual', ''), SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM), numfmt='0%')
    _set(ws, 'F15', fg.get('rps_goal', ''), SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM), numfmt='0.00%')
    _set(ws, 'F16', fg.get('rps_mtd', ''), SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM))
    _set(ws, 'F17', fg.get('rps_ytd', ''), SECTION_FONT_PLAIN, GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, MEDIUM))

    # ---- Site Inspections ----
    _set(ws, 'A18', 'Site Inspections', SECTION_FONT, GREEN_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, THIN))
    _merge(ws, 'B18:I18')
    _set(ws, 'B18', None, None, GREEN_FILL, None, _border(None, MEDIUM, MEDIUM, THIN))
    headers19 = ['Time', 'Guest', 'Position / Company', None, None, None, None, 'Sales Contact']
    _set(ws, 'A19', 'Time', TABLE_HEADER_FONT, GREEN_FILL, CENTER, _border(MEDIUM, THIN, THIN, THIN))
    _set(ws, 'B19', 'Guest', TABLE_HEADER_FONT, GREEN_FILL, CENTER, _border(THIN, THIN, THIN, THIN))
    _merge(ws, 'C19:G19')
    _set(ws, 'C19', 'Position / Company', TABLE_HEADER_FONT, GREEN_FILL, CENTER, _border(THIN, THIN, THIN, THIN))
    _merge(ws, 'H19:I19')
    _set(ws, 'H19', 'Sales Contact', TABLE_HEADER_FONT, GREEN_FILL, CENTER, _border(THIN, MEDIUM, THIN, THIN))

    row = 20
    inspections = data.get('site_inspections', []) or [{}]
    for insp in inspections:
        b = _border(THIN, THIN, THIN, THIN)
        _set(ws, f'A{row}', insp.get('time', ''), DATA_FONT, None, CENTER, b)
        _set(ws, f'B{row}', insp.get('guest', ''), DATA_FONT, None, CENTER, b)
        _merge(ws, f'C{row}:G{row}')
        _set(ws, f'C{row}', insp.get('position_company', ''), DATA_FONT, None, CENTER, b)
        for c in 'DEFG':
            _set(ws, f'{c}{row}', None, DATA_FONT, None, CENTER, b)
        _merge(ws, f'H{row}:I{row}')
        _set(ws, f'H{row}', insp.get('sales_contact', ''), DATA_FONT, None, CENTER, b)
        _set(ws, f'I{row}', None, DATA_FONT, None, CENTER, b)
        row += 1
    row += 1

    # ---- VIP Arrivals ----
    _section_header(ws, row, 'VIP Arrivals')
    row += 1
    _vip_table_header(ws, row, 'ETA', 'Remarks', has_photo_col=True)
    row += 1
    for g in data.get('vip_arrivals', []):
        fill_c, font_c = rules.color_for_code(g.get('code'))
        _vip_row(ws, row, g.get('guest', ''), g.get('code', ''), g.get('room', ''),
                  g.get('eta', ''), g.get('remarks', ''), g.get('company', ''), fill_c, font_c,
                  photo_fs_path=_photo_fs_path(g.get('photo')), has_photo_col=True)
        row += 1
    row += 1

    # ---- VIP In-House Guests ----
    _section_header(ws, row, 'VIP In-House Guests')
    row += 1
    _vip_table_header(ws, row, 'Departure day', 'Remarks', has_photo_col=True)
    row += 1
    for g in data.get('vip_inhouse', []):
        fill_c, font_c = rules.color_for_code(g.get('code'))
        _vip_row(ws, row, g.get('guest', ''), g.get('code', ''), g.get('room', ''),
                  g.get('departure_day', ''), g.get('remarks', ''), g.get('company', ''), fill_c, font_c,
                  photo_fs_path=_photo_fs_path(g.get('photo')), has_photo_col=True)
        row += 1
    row += 1

    # ---- VIP Departures ----
    _section_header(ws, row, 'VIP Departures')
    row += 1
    _vip_table_header(ws, row, 'Departure day', 'Departure Time')
    row += 1
    for g in data.get('vip_departures', []):
        fill_c, font_c = rules.color_for_code(g.get('code'))
        _vip_row(ws, row, g.get('guest', ''), g.get('code', ''), g.get('room', ''),
                  g.get('departure_day', ''), g.get('departure_time', ''), g.get('company', ''), fill_c, font_c)
        row += 1

    widths = {'A': 26, 'B': 8, 'C': 10, 'D': 16, 'E': 14, 'F': 12, 'G': 16, 'H': 14, 'I': 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.print_area = f'A1:I{row + 1}'
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # =======================================================================
    # Sheet 2: F&B Section
    # =======================================================================
    fb = wb.create_sheet('F&B Section')
    _merge(fb, 'A2:I2')
    _set(fb, 'A2', title_text, Font(name=FONT_NAME, size=16, bold=True), GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, THIN))
    _merge(fb, 'A3:I3')
    _set(fb, 'A3', quote_text, Font(name=FONT_NAME, size=16, bold=True), GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, THIN, THIN))
    _merge(fb, 'A4:I4')
    _set(fb, 'A4', 'FOOD & BEVERAGE', Font(name=FONT_NAME, size=16), GREY_FILL, CENTER, _border(MEDIUM, MEDIUM, THIN, MEDIUM))

    _merge(fb, 'A5:I5')
    _set(fb, 'A5', 'Events', SECTION_FONT, GREEN_FILL, CENTER, _border(MEDIUM, MEDIUM, MEDIUM, THIN))
    _set(fb, 'A6', 'Time', TABLE_HEADER_FONT, GREEN_FILL, CENTER, _border(MEDIUM, THIN, THIN, THIN))
    _merge(fb, 'B6:C6')
    _set(fb, 'B6', 'Meeting Room', TABLE_HEADER_FONT, GREEN_FILL, CENTER, _border(THIN, THIN, THIN, THIN))
    _merge(fb, 'D6:F6')
    _set(fb, 'D6', 'Company', TABLE_HEADER_FONT, GREEN_FILL, CENTER, _border(THIN, THIN, THIN, THIN))
    _merge(fb, 'G6:I6')
    _set(fb, 'G6', 'Sales contact', TABLE_HEADER_FONT, GREEN_FILL, CENTER, _border(THIN, MEDIUM, THIN, THIN))

    erow = 7
    events = data.get('events', []) or [{}, {}]
    for ev in events:
        b = _border(THIN, THIN, THIN, THIN)
        _set(fb, f'A{erow}', ev.get('time', ''), DATA_FONT, GREY_FILL, CENTER, b)
        _merge(fb, f'B{erow}:C{erow}')
        _set(fb, f'B{erow}', ev.get('meeting_room', ''), DATA_FONT, GREY_FILL, CENTER, b)
        _set(fb, f'C{erow}', None, DATA_FONT, GREY_FILL, CENTER, b)
        _merge(fb, f'D{erow}:F{erow}')
        _set(fb, f'D{erow}', ev.get('company', ''), DATA_FONT, GREY_FILL, CENTER, b)
        for c in 'EF':
            _set(fb, f'{c}{erow}', None, DATA_FONT, GREY_FILL, CENTER, b)
        _merge(fb, f'G{erow}:I{erow}')
        _set(fb, f'G{erow}', ev.get('sales_contact', ''), DATA_FONT, GREY_FILL, CENTER, b)
        for c in 'HI':
            _set(fb, f'{c}{erow}', None, DATA_FONT, GREY_FILL, CENTER, b)
        erow += 1

    row9 = erow + 1
    _set(fb, f'A{row9}', 'Birthday', TABLE_HEADER_FONT, GREEN_FILL, CENTER, _border(MEDIUM, THIN, MEDIUM, THIN))
    _merge(fb, f'B{row9}:C{row9}')
    _set(fb, f'B{row9}', None, None, GREEN_FILL, None, _border(THIN, THIN, MEDIUM, THIN))
    _merge(fb, f'D{row9}:I{row9}')
    _set(fb, f'D{row9}', 'Anniversary', TABLE_HEADER_FONT, GREEN_FILL, CENTER, _border(THIN, MEDIUM, MEDIUM, THIN))

    birthdays = data.get('birthday', []) or ['']
    anniversaries = data.get('anniversary', []) or ['']
    if isinstance(birthdays, str):
        birthdays = [birthdays]
    if isinstance(anniversaries, str):
        anniversaries = [anniversaries]
    n_lines = max(len(birthdays), len(anniversaries), 1)

    row10 = row9 + 1
    for i in range(n_lines):
        r = row10 + i
        bottom = MEDIUM if i == n_lines - 1 else THIN
        _merge(fb, f'A{r}:C{r}')
        bval = birthdays[i] if i < len(birthdays) else ''
        _set(fb, f'A{r}', bval, Font(name=FONT_NAME, size=13), GREY_FILL,
             Alignment(horizontal='left', vertical='center'), _border(MEDIUM, THIN, THIN, bottom))
        for c in 'BC':
            _set(fb, f'{c}{r}', None, None, GREY_FILL, None, _border(THIN, THIN, THIN, bottom))
        _merge(fb, f'D{r}:I{r}')
        aval = anniversaries[i] if i < len(anniversaries) else ''
        _set(fb, f'D{r}', aval, Font(name=FONT_NAME, size=13), GREY_FILL,
             Alignment(horizontal='left', vertical='center'), _border(THIN, MEDIUM, THIN, bottom))
        for c in 'EFGHI':
            _set(fb, f'{c}{r}', None, None, GREY_FILL, None, _border(THIN, MEDIUM, THIN, bottom))

    row12 = row10 + n_lines + 1
    _merge(fb, f'A{row12}:I{row12}')
    _set(fb, f'A{row12}', 'Food & Beverage Performance', SECTION_FONT, GREEN_FILL, CENTER, _border(MEDIUM, MEDIUM, None, THIN))

    row13 = row12 + 1
    perf_headers = [('A', 'Outlet'), ('B', 'Revenue'), ('C', 'G\u0130H'),
                    ('D', 'External guests'), ('E', 'Special promotions'), ('G', 'External guests')]
    for col, label in perf_headers:
        span = None
        if col == 'E':
            span = 'F'
        if span:
            _merge(fb, f'{col}{row13}:{span}{row13}')
        b = _border(THIN if col != 'A' else MEDIUM, MEDIUM if col == 'G' else THIN, THIN, THIN)
        _set(fb, f'{col}{row13}', label, TABLE_HEADER_FONT, None, CENTER, b)
        if span:
            _set(fb, f'{span}{row13}', None, None, None, None, b)
    _merge(fb, f'G{row13}:I{row13}')
    for c in 'HI':
        _set(fb, f'{c}{row13}', None, None, None, None, _border(THIN, MEDIUM, THIN, THIN))

    prow = row13 + 1
    for perf in data.get('fb_performance', []):
        b = _border(THIN, THIN, THIN, THIN)
        _set(fb, f'A{prow}', perf.get('outlet', ''), Font(name=FONT_NAME, size=16, bold=True), None, Alignment(horizontal='left', vertical='center', wrap_text=True), b)
        _set(fb, f'B{prow}', perf.get('revenue', None), DATA_FONT, GREY_FILL, Alignment(horizontal='right'), b, numfmt='#,##0.00')
        _set(fb, f'C{prow}', perf.get('gih', None), DATA_FONT, GREY_FILL, Alignment(horizontal='right'), b)
        _set(fb, f'D{prow}', perf.get('external_guests', None), DATA_FONT, GREY_FILL, CENTER, b)
        _merge(fb, f'E{prow}:F{prow}')
        _set(fb, f'E{prow}', perf.get('special_promotions', ''), DATA_FONT, None, CENTER, b)
        _set(fb, f'F{prow}', None, DATA_FONT, None, CENTER, b)
        _merge(fb, f'G{prow}:I{prow}')
        _set(fb, f'G{prow}', perf.get('external_guests_2', ''), DATA_FONT, None, CENTER, b)
        for c in 'HI':
            _set(fb, f'{c}{prow}', None, DATA_FONT, None, CENTER, b)
        prow += 1

    fb_widths = {'A': 22, 'B': 12, 'C': 8, 'D': 14, 'E': 12, 'F': 10, 'G': 14, 'H': 10, 'I': 10}
    for col, w in fb_widths.items():
        fb.column_dimensions[col].width = w

    fb.print_area = f'A1:I{prow + 1}'
    fb.page_setup.paperSize = 9
    fb.page_setup.orientation = 'portrait'
    fb.page_setup.fitToWidth = 1
    fb.page_setup.fitToHeight = 0
    fb.sheet_properties.pageSetUpPr.fitToPage = True

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

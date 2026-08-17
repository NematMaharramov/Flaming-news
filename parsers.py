"""
Flaming News - PDF Parsers
Parses the 4 Fairmont PMS report PDFs into structured Python dicts.
"""
import re
import pdfplumber

VIP_CODES = {'T3', 'T4', 'T5', 'T6', 'DV', 'SA', 'V1'}


def _extract_text(file_stream):
    """Return list of raw text lines across all pages of a PDF."""
    lines = []
    with pdfplumber.open(file_stream) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines.extend(text.split('\n'))
    return lines


# ---------------------------------------------------------------------------
# 1. History and Forecast -> 7 Day Forecast section
# ---------------------------------------------------------------------------
# Real "Flaming_Forecast" export column order (confirmed against a live
# sample): Date Weekday | Total Occ. | Arr. Rooms | Comp. Rooms | House Use |
# Deduct Indiv. | Deduct Group | Occ.% | Room Revenue | Average Rate |
# Dep. Rooms | Day Use Rooms | No Show Rooms | OOO Rooms | Adl. & Chl.
_FORECAST_RE = re.compile(
    r'^(\d{2}\.\d{2}\.\d{2})\s+(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+'
    r'(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+'
    r'([\d.]+)%\s+([\d,]+\.\d+)\s+([\d.]+)\s+'
    r'(\d+)\s+.*$'
)


def _to_ddmmyyyy(short_date):
    """'16.08.26' -> '16.08.2026'"""
    dd, mm, yy = short_date.split('.')
    return f'{dd}.{mm}.20{yy}'


def parse_forecast(file_stream):
    """
    Parse the Flaming_Forecast PDF into the 7-day forecast structure:
    {'dates': [...7], 'occupancy_pct': [...7], 'rooms_occupied': [...7],
     'adr': [...7], 'arrivals': [...7], 'departures': [...7]}
    Only the first 7 matching daily rows are used (Subtotal/Total lines
    don't match the date-anchored pattern, so they're naturally skipped).
    """
    lines = _extract_text(file_stream)
    days = []
    for line in lines:
        m = _FORECAST_RE.match(line.strip())
        if not m:
            continue
        (date, total_occ, arr_rooms, comp_rooms, house_use, deduct_indiv,
         deduct_group, occ_pct, room_revenue, avg_rate, dep_rooms) = m.groups()
        days.append({
            'date': _to_ddmmyyyy(date),
            'occupancy_pct': round(float(occ_pct) / 100, 4),
            'rooms_occupied': int(total_occ),
            'adr': float(avg_rate),
            'arrivals': int(arr_rooms),
            'departures': int(dep_rooms),
        })
        if len(days) == 7:
            break

    return {
        'dates': [d['date'] for d in days],
        'occupancy_pct': [d['occupancy_pct'] for d in days],
        'rooms_occupied': [d['rooms_occupied'] for d in days],
        'adr': [d['adr'] for d in days],
        'arrivals': [d['arrivals'] for d in days],
        'departures': [d['departures'] for d in days],
    }


# ---------------------------------------------------------------------------
# 2. Guests INH - By Room -> VIP In-House Guests section
# ---------------------------------------------------------------------------
def parse_in_house(file_stream):
    lines = _extract_text(file_stream)
    records = []
    current = None
    main_re = re.compile(
        r'^(\d{4}(?:\s*,\s*\d{4})*)\s+([A-Za-z][^0-9]+?)\s+'
        r'(?:(?:T-|C-|S-)\s*(.+?)\s+)?'
        r'(\d{2}\.\d{2}\.\d{2})(\d{2}\.\d{2}\.\d{2})([A-Z0-9]+)\s+(\d+)\s+(\d+)\s*([A-Z]{2})\s+(\S+)?$'
    )
    skip_words = ('VIP', 'SA', 'V1', 'DV', 'T3', 'T4', 'T5', 'T6', 'Accompanying', 'Res.')
    for i, raw in enumerate(lines):
        line = raw.rstrip()
        m = main_re.match(line)
        if m:
            room, name, company, arr, dep, roomtype, adl, chl, paymth, rate = m.groups()
            company = (company or '').strip()
            if i + 1 < len(lines):
                nxt = lines[i + 1].strip()
                if (re.match(r'^[A-Za-z .()]+$', nxt) and
                        not any(nxt.startswith(w) or nxt == w for w in skip_words)):
                    company = (company + ' ' + nxt).strip()
            current = {
                'room': room.split(',')[0].strip(),
                'name': name.strip(),
                'company': company,
                'arr_date': _to_ddmmyyyy(arr),
                'dep_date': _to_ddmmyyyy(dep),
                'vip_code': None,
                'raw_specials': [],
            }
            records.append(current)
            continue
        if current and re.match(r'^(T3|T4|T5|T6|DV|SA|V1)$', line.strip()):
            current['vip_code'] = line.strip()
        if current and line.strip().startswith('Specials:'):
            current['raw_specials'].append(line.strip().replace('Specials:', '').strip())
    return [r for r in records if r['vip_code']]


# ---------------------------------------------------------------------------
# 3. Arrivals: Detailed -> VIP Arrivals section
# ---------------------------------------------------------------------------
def parse_arrivals(file_stream):
    lines = _extract_text(file_stream)
    records = []
    current = None
    last_room = ''
    main_re = re.compile(
        r'^(\d{4})?\s*([A-Za-z][^0-9]+?)\s+'
        r'(?:(?:T-|C-|S-)\s*(.+?)\s+)?'
        r'(\d{2}\.\d{2}\.\d{2})\s+(\d{2}\.\d{2}\.\d{2})\s+([A-Z0-9]+)\s+'
        r'(\d+)\s+(\d+)\s+(\d+)\s+\S+\s+\S+\s+\S+\s+(\S+)\s+[A-Z]{3}\s+[\d.,]+\S*\s+[\d.]+$'
    )
    vip_line_re = re.compile(r'^(\d+)\s+(T3|T4|T5|T6|DV|SA|V1)(?:\s+(\d{1,2}:\d{2}))?')
    skip_words = ('VIP', 'SA', 'V1', 'DV', 'T3', 'T4', 'T5', 'T6', 'Accompanying',
                  'Reservation', 'RESERVATIONS', 'CASHIER', 'Specials', 'Membership',
                  'Traces', 'Routing')
    for i, raw in enumerate(lines):
        line = raw.rstrip()
        m = main_re.match(line)
        if m:
            room, name, company, arr, dep, roomtype, adl, chl, rms, rate = m.groups()
            company = (company or '').strip()
            same_guest = bool(records) and records[-1]['name'].strip() == name.strip()
            if room:
                last_room = room
                room_val = room
            elif same_guest:
                room_val = last_room
            else:
                room_val = ''  # different guest, room unreadable in source PDF -> leave blank/flag
            if i + 1 < len(lines):
                nxt = lines[i + 1].strip()
                if (re.match(r'^[A-Za-z .()]+$', nxt) and
                        not any(nxt.startswith(w) for w in skip_words) and
                        not vip_line_re.match(nxt)):
                    company = (company + ' ' + nxt).strip()
            current = {
                'room': room_val,
                'name': name.strip(),
                'company': company,
                'arr_date': _to_ddmmyyyy(arr),
                'eta': '',
                'vip_code': None,
                'raw_specials': [],
            }
            records.append(current)
            continue
        vm = vip_line_re.match(line.strip())
        if current and vm:
            current['vip_code'] = vm.group(2)
            current['eta'] = vm.group(3) or ''
        if current and line.strip().startswith('Specials:'):
            current['raw_specials'].append(line.strip().replace('Specials:', '').strip())
    return [r for r in records if r['vip_code']]


# ---------------------------------------------------------------------------
# 4. Departures -> VIP Departures section
# ---------------------------------------------------------------------------
# Real-world "Departures" PMS exports are messier than the other reports:
# name/company sometimes wrap onto the next physical line mid-word, and the
# numeric columns after Rms (Nts/RoomType/BlockCode/RateCode/ResStatus/
# DepTime/PayMth/Balance) don't keep a fixed token count row to row (some
# columns are blank). Since we only need room/name/company/vip_code/
# dep_date/dep_time for Flaming News, the parser captures a generous "rest
# of line" tail after the VIP code + dates and pulls dep_time out of that
# with a separate search, rather than trying to positionally match every
# trailing column.
_DEP_MAIN_RE = re.compile(
    r'^(\d{3,4})\s+([A-Za-z][^0-9]+?)\s+'
    r'(?:(?:T-|C-|S-)\s*(.+?)\s+)?'
    r'(T3|T4|T5|T6|DV|SA|V1)\s+'
    r'(\d{2}\.\d{2}\.\d{2})\s+(\d{2}\.\d{2}\.\d{2})\s+(.*)$'
)
_DEP_TIME_RE = re.compile(r'\b(\d{1,2}:\d{2})\b')
_TITLE_SUFFIX_RE = re.compile(r',\s*(MR|MRS|MS|MISS|DOC|JR)\.?\s*$', re.I)
_DEP_SKIP_WORDS = ('G-', 'Res.Comments', 'Reservation', 'General', 'Specials',
                    'Membership', 'Profile', 'Page', 'Filter', 'Room Class',
                    'Departure Time', 'Profile Type')


def parse_departures(file_stream):
    lines = _extract_text(file_stream)
    records = []
    current = None
    for i, raw in enumerate(lines):
        line = raw.rstrip()
        m = _DEP_MAIN_RE.match(line)
        if m:
            room, name, company, vip_code, arr, dep, rest = m.groups()
            name = name.strip()
            company = (company or '').strip()
            dep_time_match = _DEP_TIME_RE.search(rest)
            dep_time = dep_time_match.group(1) if dep_time_match else ''

            # Handle the common one-line wrap: a name/company fragment that
            # spilled onto the very next physical line.
            if i + 1 < len(lines):
                nxt = lines[i + 1].strip()
                is_skippable = any(nxt.startswith(w) for w in _DEP_SKIP_WORDS)
                looks_textual = bool(re.match(r'^\(?[A-Za-z ,.()]+\)?$', nxt))
                has_upper = any(ch.isupper() for ch in nxt)
                if looks_textual and not is_skippable and has_upper:
                    if not _TITLE_SUFFIX_RE.search(name) and re.match(r'^[A-Za-z]{1,2}$', nxt):
                        name = name + nxt  # e.g. "Ibrahim,Muhammad,M" + "R"
                    elif not _TITLE_SUFFIX_RE.search(name):
                        name = (name + ' ' + nxt).strip()  # wrapped full name
                    else:
                        company = (company + ' ' + nxt).strip()  # wrapped company

            current = {
                'room': room,
                'name': name,
                'company': company,
                'vip_code': vip_code,
                'dep_date': _to_ddmmyyyy(dep),
                'dep_time': dep_time,
                'raw_specials': [],
            }
            records.append(current)
            continue
        if current and line.strip().startswith('Specials:'):
            current['raw_specials'].append(
                line.strip().replace('Specials:', '').strip())
    return records


# ---------------------------------------------------------------------------
# Clean company prefixes (T-/C-/S-) — used post-parse
# ---------------------------------------------------------------------------
def clean_company(name):
    if not name:
        return ''
    return re.sub(r'^(T-|C-|S-)\s*', '', name).strip()


# ---------------------------------------------------------------------------
# 5. Restaurant / F&B outlet reports -> Food & Beverage Performance section
# ---------------------------------------------------------------------------
# The daily "Outlet revenues" export is an .xlsx with one row per outlet:
#   Outlet name | In house guests | Outside guests | Total PAX | Revenue | Reservation
# Outlet names in that export are sometimes abbreviated versus the Flaming
# News template's labels (e.g. "Bar19" vs "Bar 19", "H2O" vs "H2O Pool Bar",
# "BQT" vs "Banquet") — normalise them to the template's canonical names so
# they merge into the right row automatically.
OUTLET_NAME_MAP = {
    'nur lounge': 'Nur Lounge',
    'le bistro': 'Le Bistro',
    'bar19': 'Bar 19',
    'bar 19': 'Bar 19',
    'jazz club': 'Jazz Club',
    'balcon cafe': 'Balcon Café',
    'balcon café': 'Balcon Café',
    'minibar': 'Minibar',
    'in-room dining': 'In-Room Dining',
    'in room dining': 'In-Room Dining',
    'h2o': 'H2O Pool Bar',
    'h2o pool bar': 'H2O Pool Bar',
    'bqt': 'Banquet',
    'banquet': 'Banquet',
}


def _canonical_outlet(name):
    key = (name or '').strip().lower()
    return OUTLET_NAME_MAP.get(key, (name or '').strip())


def parse_fb_report_xlsx(file_stream):
    """
    Parse the daily "Outlet revenues" F&B export (.xlsx) into per-outlet
    Revenue / GIH (in-house guest covers) / External guest covers, matched
    against the template's outlet names.

    Expected columns (header row, any exact casing/spacing tolerated by
    fuzzy header matching): Outlet name, In house guests, Outside guests,
    Total PAX, Revenue, Reservation.

    Returns: list of {'outlet': str, 'revenue': float|None, 'gih': int|None,
                       'external_guests': int|None}
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_stream, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

    def _col(*candidates):
        for cand in candidates:
            for i, h in enumerate(header):
                if cand in h:
                    return i
        return None

    col_outlet = _col('outlet')
    col_gih = _col('in house', 'in-house', 'inhouse')
    col_external = _col('outside', 'external')
    col_revenue = _col('revenue')

    results = []
    for row in rows[1:]:
        if col_outlet is None or col_outlet >= len(row):
            continue
        outlet_raw = row[col_outlet]
        if not outlet_raw or not str(outlet_raw).strip():
            continue

        def _get(i):
            if i is None or i >= len(row):
                return None
            v = row[i]
            return v if v not in (None, '') else None

        results.append({
            'outlet': _canonical_outlet(str(outlet_raw)),
            'revenue': _get(col_revenue),
            'gih': _get(col_gih),
            'external_guests': _get(col_external),
        })
    return results


def parse_fb_report_pdf(file_stream):
    """
    Fallback for a restaurant / F&B report supplied as a PDF instead of the
    structured "Outlet revenues" Excel export. Looks, per known outlet name,
    for the nearest run of numbers on the same or next line and takes the
    first three as (revenue, gih_covers, external_covers). Anything it
    cannot confidently read is left blank rather than guessed.
    """
    lines = _extract_text(file_stream)
    results = []
    for idx, raw in enumerate(lines):
        line = raw.strip()
        matched_outlet = None
        for key, canonical in OUTLET_NAME_MAP.items():
            if key in line.lower():
                matched_outlet = canonical
                break
        if not matched_outlet or any(r['outlet'] == matched_outlet for r in results):
            continue

        nums = re.findall(_NUM, line)
        if len(nums) < 2 and idx + 1 < len(lines):
            nums += re.findall(_NUM, lines[idx + 1].strip())

        def _num(i, cast=float):
            if i < len(nums):
                try:
                    return cast(nums[i].replace(',', ''))
                except ValueError:
                    return None
            return None

        results.append({
            'outlet': matched_outlet,
            'revenue': _num(0, float),
            'gih': _num(1, int),
            'external_guests': _num(2, int),
        })
    return results


_NUM = r'[\d,]+(?:\.\d+)?'


def parse_fb_report(file_stream, filename=''):
    """Dispatch to the xlsx parser (the real daily report format) or the PDF
    heuristic fallback, based on the uploaded file's extension."""
    if filename.lower().endswith(('.xlsx', '.xlsm', '.xls')):
        return parse_fb_report_xlsx(file_stream)
    return parse_fb_report_pdf(file_stream)
